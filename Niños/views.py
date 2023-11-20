from django.shortcuts import render
from django.http import HttpRequest, HttpResponse
from .forms import FormularioNiños, FormularioExportarExcel
from .models import Niños_tabla2
from crispy_forms.helper import FormHelper
from crispy_forms.layout import Layout, Field
from django.contrib import messages
from datetime import datetime
import openpyxl

# Create your views here.
"""def app1_vista(request): 
    return render(request,'niños.html') 
""" 
class FormularioNiñosView(HttpRequest): 
    def init(self, args, **kwargs):
        super(FormularioNiños, self).init(args, **kwargs)
        self.helper = FormHelper()
        self.helper.layout = Layout(Field('nombre', css_class='form-control-small'))

    def index(request):
        Niños=FormularioNiños() 
        return render (request,"NiñosIndex.html",{"form":Niños}) 
    
    def procesar_form(request): 
        Niños=FormularioNiños(request.POST) 
        if Niños.is_valid(): 
            Niños.save() 
            Niños=FormularioNiños() 
        return render(request, "NiñosIndex.html",{"form":Niños, "mensaje":'OK'}) 
    
    def listar_Niños(request): 
        Niños=Niños_tabla2.objects.all() 
        return render(request, "NiñosLista.html",{"Niños":Niños})  
    
    def editar(request, id_niño):  
        Niños_edit=Niños_tabla2.objects.filter(id=id_niño).first() 
        form=FormularioNiños(instance=Niños_edit) 
        return render (request,"NiñoEdit.html",{"form":form, 'Niños':Niños_edit}) 
    
    def actualizar_niño(request, id_niño):
        Niño = Niños_tabla2.objects.get(pk=id_niño)
        form= FormularioNiños(request.POST, instance=Niño)
        if form.is_valid():
            form.save() 
        messages.success(request,'Niño actualizado correctamente')
        Niños_actualizar = Niños_tabla2.objects.all() 

        return render(request, "NiñosLista.html",{"form":form, "Niños":Niños_actualizar}) 
    
    def delete(request,id_niño): 
        Niño=Niños_tabla2.objects.get(pk=id_niño) 
        Niño.delete()
        messages.success(request,'Niño eliminado correctamente')
        Niños=Niños_tabla2.objects.all() 
        return render(request,"NiñosLista.html",{"Niños":Niños})
    
    
    def export_excel_ninos(request):
        form = FormularioExportarExcel(request.POST or None)
        if request.method == 'POST' and form.is_valid():
            option = form.cleaned_data.get('option')
            mes = form.cleaned_data.get('mes')
            if option == 'mes' and mes is not None:
                queryset = Niños_tabla2.objects.filter(FechaIngreso__year=mes.year, FechaIngreso__month=mes.month)
            else:
                queryset = Niños_tabla2.objects.all()
            date = datetime.now().strftime('%d-%m-%Y')
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Niños_{date}.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Niños'

            # Write header row
            header = ['Nombre', 'Fecha_nacimiento', 'Peso', 'Edad', 'Talla', 'Fecha_ingreso']
            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            # Write data rows
            queryset = Niños_tabla2.objects.all().values_list('nombre', 'fechaDeNacimiento', 'peso',  'edad', 'infoTalla', 'FechaIngreso')
            for row_num, row in enumerate(queryset, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num+1, column=col_num)
                    cell.value = cell_value
                

            
            for col_letter in ["B", "F"]:
                worksheet.column_dimensions[col_letter].auto_size = True
            
            workbook.save(response)
            
            return response
        
        else:
            return render(request, "exportarExcelNiños.html", {"form": form})

