from django.shortcuts import render
from django.shortcuts import render
from django.http import HttpRequest, HttpResponse
from .forms import FormularioComidas 
from .forms import FormularioExportarExcel
from .models import comida
from django.contrib import messages
from datetime import datetime
import openpyxl


class FormularioComidasView(HttpRequest):  
    def index(request):
        Comidas=FormularioComidas() 
        return render (request,"ComidasIndex.html",{"form":Comidas}) 
    
    def procesar_form(request): 
        Comidas=FormularioComidas(request.POST) 
        if Comidas.is_valid(): 
            Comidas.save() 
            Comidas=FormularioComidas() 
        return render(request, "ComidasIndex.html",{"form":Comidas, "mensaje":'OK'}) 
    
    def listar_Comidas(request): 
        Comidas=comida.objects.all() 
        return render(request, "ComLista.html",{"Comidas":Comidas}) 
    
    
    def editar(request, id_comida):  
        Comidas_edit=comida.objects.filter(id=id_comida).first() 
        form=FormularioComidas(instance=Comidas_edit) 
        return render (request,"ComidaEdit.html",{"form":form, 'Comidas':Comidas_edit}) 
    
    def actualizar_comida(request, id_comida):
        Comida= comida.objects.get(pk=id_comida)
        form= FormularioComidas(request.POST, instance=Comida)
        if form.is_valid():
            form.save() 
        messages.success(request,'Comida actualizada correctamente')
        Comidas = comida.objects.all()
        return render(request, "ComLista.html",{"form":form, "Comidas":Comidas}) 

    def delete(request,id_comida): 
        Comida=comida.objects.get(pk=id_comida) 
        Comida.delete()
        messages.success(request,'Comida eliminada correctamente')
        Comidas=comida.objects.all() 
        return render(request,"ComLista.html",{"Comidas":Comidas})
    
    def export_excel(request):
        form = FormularioExportarExcel(request.POST or None)
        if request.method == 'POST' and form.is_valid():
            option = form.cleaned_data.get('option')
            mes = form.cleaned_data.get('mes')
            if option == 'mes' and mes is not None:
                queryset = comida.objects.filter(FechaIngreso__year=mes.year, FechaIngreso__month=mes.month)
            else:
                queryset = comida.objects.all()

            date = datetime.now().strftime('%d-%m-%Y')
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Productos_{date}.xlsx"'

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Productos'

            # Write header row
            header = ['Producto',  'Modo de Ingreso', 'Peso',  'Fecha de Ingreso', 'Modo de Salida', 'Fecha de Salida']
            for col_num, column_title in enumerate(header, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = column_title

            # Write data rows
            queryset = queryset.values_list('producto',  'modoIngreso',  'peso', 'FechaIngreso', 'modoSalida', 'fechaSalida')
            for row_num, row in enumerate(queryset, 1):
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num+1, column=col_num)
                    cell.value = cell_value

            for col_letter in ["D", "F"]:
                worksheet.column_dimensions[col_letter].auto_size = True
            
            workbook.save(response)

            return response
        else:
            return render(request, "ExportarExcel.html", {"form": form})
